from openpyxl import load_workbook
from docx import Document
import os
import win32com.client
import time
moban_path = 'D:\\pdf\\'
moban_file='D:\\1.docx'
mingdan_file='D:\\1.xlsx'
zhanwei='[XX]'


def wps_to_pdf(word_path, pdf_path=None):
    """
    稳定版：基于WPS Office实现Word转PDF（解决PDF损坏问题）
    :param word_path: Word文件完整路径（支持.doc/.docx）
    :param pdf_path: 输出PDF路径，默认同目录同名
    """
    # 1. 基础校验
    if not os.path.exists(word_path):
        raise FileNotFoundError(f"Word文件不存在：{word_path}")
    # 处理输出路径，避免特殊字符/空格问题（路径用双引号包裹）
    if pdf_path is None:
        pdf_path = os.path.splitext(word_path)[0] + ".pdf"
    # 确保输出目录存在
    pdf_dir = os.path.dirname(pdf_path)
    if not os.path.exists(pdf_dir):
        os.makedirs(pdf_dir)

    wps_app = None
    try:
        # 2. 启动WPS（后台模式）
        wps_app = win32com.client.Dispatch("KWps.Application")
        wps_app.Visible = False
        wps_app.DisplayAlerts = False  # 禁用所有弹窗

        # 3. 打开文档（取消只读模式，避免导出权限问题）
        doc = wps_app.Documents.Open(
            FileName=word_path,
            ReadOnly=False,  # 关键：只读模式可能导致导出失败
            ConfirmConversions=False,
            AddToRecentFiles=False  # 不加入最近文件列表
        )

        # 4. 核心修复：用ExportAsFixedFormat导出PDF（替代SaveAs）
        # ExportFormat=17 是PDF格式（和Word的wdExportFormatPDF一致，WPS兼容）
        doc.ExportAsFixedFormat(
            OutputFileName=pdf_path,
            ExportFormat=17,  # 17=PDF，24=XPS（WPS/Word通用常量）
            OpenAfterExport=False,  # 导出后不打开PDF
            OptimizeFor=0,  # 0=屏幕优化，1=打印优化
            Range=0,  # 0=全部内容，1=当前页，2=指定范围
            Item=0,  # 0=文档内容，1=批注，2=文档和批注
            IncludeDocProps=True,  # 包含文档属性
            KeepIRM=True  # 保留权限设置
        )

        # 5. 关键延迟：确保WPS完全写入文件（避免进程提前终止）
        time.sleep(1)  # 大文件可适当增加（如2-3秒）

        print(f"✅ 转换成功！PDF路径：{pdf_path}")
        # 验证文件是否有效（检查文件大小）
        if os.path.getsize(pdf_path) < 100:  # 有效PDF至少几百字节
            raise RuntimeError("PDF文件过小，可能导出不完整")

    except Exception as e:
        # 出错时删除损坏的PDF文件
        if os.path.exists(pdf_path):
            os.remove(pdf_path)
        raise RuntimeError(f"❌ 转换失败：{str(e)}")
    finally:
        # 6. 安全清理进程（确保文档关闭后再退出WPS）
        if 'doc' in locals():
            doc.Close(SaveChanges=False)
            del doc  # 释放文档对象
        if wps_app is not None:
            wps_app.Quit()
            del wps_app  # 释放WPS对象
        # 额外延迟，确保进程完全退出
        time.sleep(0.5)


def convert(moban_path,moban_file,mingdan_file,zhanwei,n):

    # 加载Excel工作簿
    wb = load_workbook(mingdan_file)
    # 选择活动的工作表或通过名称选择特定的工作表
    sheet = wb.active  # 获取活动工作表
    # 或 sheet = wb['Sheet1']  # 通过名称获取特定工作表
    # 或者，如果你想获取所有数据到一个列表中：
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row[n])
    print("要处理的文件数量是：",len(data))

    #批量生成word
    for name in data:
        # 如果存在空值，跳过
        if name == None:
            continue
        # 加载文档
        doc = Document(moban_file)
        print('当前处理的是：',name)
        # 遍历所有段落，找到需要替换的文本并替换
        for paragraph in doc.paragraphs:
            if zhanwei in paragraph.text:
                inline = paragraph.runs
                for i in range(len(inline)):
                    if zhanwei in inline[i].text:
                        inline[i].text = inline[i].text.replace(zhanwei, name)
        # 保存文档
        word_tar = moban_path + '/' + name +'.docx'
        doc.save(word_tar)
        
        wps_to_pdf(word_tar)


    print("处理完成")



# if __name__ == '__main__':
#     convert(moban_path,moban_file,mingdan_file,zhanwei)